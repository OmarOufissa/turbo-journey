#!/usr/bin/env python3
"""
Thorough duplicate-matricule analysis.

For every page flagged as `duplicate_matricule` in reports/process-report.json,
compare it against the "first" certificate page that already claimed the same
matricule. Uses the official employees.xlsx (canonical Nom & Prénom + N° du
titre per matricule) as ground truth.

For each pair, OCR full text of both pages and decide:
  - TRUE_DUPLICATE        : same employee (name matches canonical), same N° de
                            titre on both pages -> genuinely the same document.
  - DIFFERENT_CERTIFICATE : same employee, but different N° de titre -> a
                            second/additional certificate for that employee.
  - REQUIRES_MANUAL_REVIEW: anything else (name mismatch suggests the
                            "duplicate" page actually belongs to a different
                            employee whose matricule was misread, or OCR too
                            poor to confirm).

Conservative by design: never auto-classify as TRUE_DUPLICATE unless both the
canonical employee name AND the N° de titre clearly match on both pages.
"""
import json
import os
import re
import subprocess
import tempfile
import unicodedata
from pathlib import Path

BASE_DIR = Path(__file__).parent.parent
RAW_DIR = BASE_DIR / "raw-pdfs"
REPORT_PATH = BASE_DIR / "uploads" / "pdfs" / "process-report.json"
EXCEL_PATH = BASE_DIR / "server" / "seeds" / "data" / "employees.xlsx"
OUT_PATH = BASE_DIR / "reports" / "duplicate-analysis.json"

OCR_CACHE_PATH = BASE_DIR / "reports" / ".dup-ocr-cache.json"


def normalize_text(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.upper()


def name_tokens(name: str) -> list[str]:
    norm = normalize_text(name)
    return [t for t in re.split(r"[^A-Z]+", norm) if len(t) >= 2]


def name_match_score(canonical_name: str, page_text: str) -> tuple[int, int]:
    """Return (tokens_found, total_tokens) for canonical name's tokens in page_text."""
    norm_text = normalize_text(page_text)
    tokens = name_tokens(canonical_name)
    if not tokens:
        return 0, 0
    found = sum(1 for t in tokens if t in norm_text)
    return found, len(tokens)


def normalize_titre(t: str) -> str:
    t = t.upper()
    t = re.sub(r"[\s_\-/.]", "", t)
    return t


TITRE_PATTERNS = [
    re.compile(r"TITRE\s+D[’'’]?HABILITATION\s*N°?\s*[:\.]?\s*([A-Z0-9][A-Z0-9/_\- ]{3,20})", re.IGNORECASE),
    re.compile(r"N°\s*(?:DU\s*)?TITRE\s*[:\.]?\s*([A-Z0-9][A-Z0-9/_\- ]{3,20})", re.IGNORECASE),
]


def extract_titre(text: str) -> str | None:
    for pat in TITRE_PATTERNS:
        m = pat.search(text)
        if m:
            candidate = m.group(1).strip()
            # cut at common trailing words that get glommed on by OCR
            candidate = re.split(r"\s{2,}|\bNOM\b|\bMATRICULE\b", candidate, flags=re.IGNORECASE)[0]
            candidate = candidate.strip()
            if candidate:
                return candidate
    return None


# ---------------------------------------------------------------------------
# OCR with caching
# ---------------------------------------------------------------------------
_cache: dict[str, str] = {}
if OCR_CACHE_PATH.exists():
    _cache = json.loads(OCR_CACHE_PATH.read_text())


def cache_key(pdf_file: str, page_num: int) -> str:
    return f"{pdf_file}::{page_num}"


def ocr_full_page(pdf_file: str, page_num: int) -> str:
    key = cache_key(pdf_file, page_num)
    if key in _cache:
        return _cache[key]

    pdf_path = RAW_DIR / pdf_file
    text = ""
    if pdf_path.exists():
        with tempfile.TemporaryDirectory() as tmpdir:
            ppm_base = os.path.join(tmpdir, "p")
            subprocess.run(
                ["pdftoppm", "-r", "250", "-f", str(page_num), "-l", str(page_num), str(pdf_path), ppm_base],
                capture_output=True,
            )
            ppm_candidates = sorted(Path(tmpdir).glob("p*.ppm"))
            if ppm_candidates:
                result = subprocess.run(
                    ["tesseract", str(ppm_candidates[0]), "stdout", "-l", "fra", "--psm", "3"],
                    capture_output=True, text=True, timeout=120,
                )
                text = result.stdout

    _cache[key] = text
    # periodically persist
    if len(_cache) % 10 == 0:
        OCR_CACHE_PATH.write_text(json.dumps(_cache, ensure_ascii=False))
    return text


# ---------------------------------------------------------------------------
# Load Excel ground truth
# ---------------------------------------------------------------------------
def load_excel():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    by_matricule = {}
    by_titre = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        mat = row[idx["MATRICULE"]]
        if mat is None:
            continue
        mat = str(mat).strip()
        name = (row[idx["Nom & Prénom "]] or "").strip()
        titre = row[idx["N° du titre"]]
        titre = str(titre).strip() if titre else None
        by_matricule[mat] = {"name": name, "titre": titre}
        if titre:
            by_titre.setdefault(normalize_titre(titre), []).append((mat, name))

    return by_matricule, by_titre


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    report = json.loads(REPORT_PATH.read_text())
    review_queue = report["review_queue"]
    duplicates = [r for r in review_queue if r["type"] == "duplicate_matricule"]

    by_matricule, by_titre = load_excel()

    first_loc_pat = re.compile(r"from (.+) p(\d+); this page")

    results = []
    for idx_d, item in enumerate(duplicates, 1):
        claimed_mat = item["matricule"]
        dup_file = item["pdf_file"]
        dup_page = item["page_num"]

        m = first_loc_pat.search(item["reason"])
        first_file, first_page = m.group(1), int(m.group(2))

        print(f"[{idx_d}/{len(duplicates)}] mat={claimed_mat} dup={dup_file} p{dup_page} "
              f"vs first={first_file} p{first_page}")

        dup_text = ocr_full_page(dup_file, dup_page)
        first_text = ocr_full_page(first_file, first_page)

        canonical = by_matricule.get(claimed_mat)
        canonical_name = canonical["name"] if canonical else None
        canonical_titre = canonical["titre"] if canonical else None

        first_name_match = name_match_score(canonical_name, first_text) if canonical_name else (0, 0)
        dup_name_match = name_match_score(canonical_name, dup_text) if canonical_name else (0, 0)

        first_titre = extract_titre(first_text)
        dup_titre = extract_titre(dup_text)

        decision = "REQUIRES_MANUAL_REVIEW"
        reason = ""
        suggested_matricule = None

        def ratio(score):
            f, t = score
            return (f / t) if t else 0.0

        first_ok = ratio(first_name_match) >= 0.5
        dup_ok = ratio(dup_name_match) >= 0.5

        if not canonical:
            reason = f"Matricule {claimed_mat} not found in employees.xlsx — cannot verify identity"
        elif first_ok and dup_ok:
            # Both pages plausibly belong to the canonical employee
            ft = normalize_titre(first_titre) if first_titre else None
            dt = normalize_titre(dup_titre) if dup_titre else None
            if ft and dt and ft == dt:
                decision = "TRUE_DUPLICATE"
                reason = (f"Same matricule {claimed_mat} ({canonical_name}), same N° de titre "
                          f"'{first_titre}' == '{dup_titre}' on both pages — same document scanned twice.")
            elif ft and dt and ft != dt:
                decision = "DIFFERENT_CERTIFICATE"
                reason = (f"Same employee {canonical_name} (matricule {claimed_mat}), but different "
                          f"N° de titre: first='{first_titre}' vs duplicate='{dup_titre}' — "
                          f"appears to be an additional/renewal certificate, not a duplicate.")
            else:
                reason = (f"Same employee name found on both pages ({canonical_name}), but N° de titre "
                          f"could not be reliably read on one or both pages "
                          f"(first='{first_titre}', duplicate='{dup_titre}') — cannot confirm same document.")
        elif first_ok and not dup_ok:
            # The "duplicate" page likely belongs to someone else
            reason = (f"Canonical name for matricule {claimed_mat} ('{canonical_name}') matches the FIRST "
                      f"page but NOT the duplicate page — the duplicate page likely belongs to a "
                      f"different employee whose matricule was misread.")
            if dup_titre:
                hits = by_titre.get(normalize_titre(dup_titre), [])
                if hits:
                    other_mat, other_name = hits[0]
                    suggested_matricule = other_mat
                    onm, ont = name_match_score(other_name, dup_text)
                    if ratio((onm, ont)) >= 0.5:
                        reason += (f" N° de titre '{dup_titre}' on the duplicate page matches matricule "
                                   f"{other_mat} ({other_name}) in employees.xlsx, and that name appears "
                                   f"on the duplicate page — likely belongs to {other_mat}, not {claimed_mat}.")
        elif not first_ok and dup_ok:
            reason = (f"Canonical name for matricule {claimed_mat} ('{canonical_name}') matches the "
                      f"DUPLICATE page but NOT the first/kept page — the FIRST/kept certificate may "
                      f"itself be misassigned. Needs review of the kept certificate too.")
        else:
            reason = (f"Canonical name for matricule {claimed_mat} ('{canonical_name}') was not clearly "
                      f"found on either page (OCR quality) — cannot confirm identity for either page.")

        results.append({
            "claimed_matricule": claimed_mat,
            "canonical_name": canonical_name,
            "canonical_titre": canonical_titre,
            "first": {"file": first_file, "page": first_page, "titre": first_titre,
                      "name_match": f"{first_name_match[0]}/{first_name_match[1]}"},
            "duplicate": {"file": dup_file, "page": dup_page, "titre": dup_titre,
                          "name_match": f"{dup_name_match[0]}/{dup_name_match[1]}"},
            "decision": decision,
            "reason": reason,
            "suggested_matricule": suggested_matricule,
        })

    OCR_CACHE_PATH.write_text(json.dumps(_cache, ensure_ascii=False))

    from collections import Counter
    counts = Counter(r["decision"] for r in results)
    print("\n=== SUMMARY ===")
    for k, v in counts.items():
        print(f"  {k}: {v}")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps({"results": results, "summary": dict(counts)}, ensure_ascii=False, indent=2))
    print(f"\nWritten to {OUT_PATH}")


if __name__ == "__main__":
    main()
